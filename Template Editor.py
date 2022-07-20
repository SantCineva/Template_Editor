############################################
# Created by @SantCineva
############################################

import win32com.client as win32
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter


file2conv = 'E:\sant.xls'           # Locatia documentului de convertit
file_conv = 'E:\sant_conv.xlsx'     # Locatia documentului convertit


test_rez = "OK"             # Rezultatul testului [OK] , [NOK] , [NOT] , [NTB] , [NIM] , [PARTLY]  
sw_ver = "E329.2"           # Versiunea de software
variant = "Premium Plus"    # Varianta HU
test_region = "Europe"      # Regiunea
hw_rev = "D5"               # Varianta HW


col_util = [6 , 10 , 12 , 13 , 14]      # Indexul coloanei (de ex. A -> 1, B -> 2, etc.)


def convert_to_xlsx(sursa , destinatie , file_format):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(sursa)
    wb.SaveAs(destinatie , file_format)
    wb.Close()
    excel.Application.Quit()


if __name__ == "__main__":

    convert_to_xlsx(file2conv , file_conv , 51)

    wb = load_workbook(file_conv)
    ws = wb.active

    for column in col_util:

        rand = 2
        
        while (ws['A' + str(rand)].value):
            if (column == 6):
                ws[get_column_letter(column) + str(rand)].value = test_rez
                print(f"New value for > TEST RESULT < has been set to | {ws[get_column_letter(column) + str(rand)].value} | on column | {get_column_letter(column)} | rand | {str(rand)} |")

            if (column == 10):
                ws[get_column_letter(column) + str(rand)].value = sw_ver
                print(f"New value for > SW VERSION < has been set to | {ws[get_column_letter(column) + str(rand)].value} | on column | {get_column_letter(column)} | rand | {str(rand)} |")

            if (column == 12):
                ws[get_column_letter(column) + str(rand)].value = variant
                print(f"New value for > VARIANT < has been set to | {ws[get_column_letter(column) + str(rand)].value} | on column | {get_column_letter(column)} | rand | {str(rand)} |")

            if (column == 13):
                ws[get_column_letter(column) + str(rand)].value = test_region
                print(f"New value for > TEST REGION < has been set to | {ws[get_column_letter(column) + str(rand)].value} | on column | {get_column_letter(column)} | rand | {str(rand)} |")

            if (column == 14):
                ws[get_column_letter(column) + str(rand)].value = hw_rev
                print(f"New value for > HW VERSION < has been set to | {ws[get_column_letter(column) + str(rand)].value} | on column | {get_column_letter(column)} | rand | {str(rand)} |")

            rand += 1

    wb.save(file_conv)

    print("""   ______                    __     _____                                     
  / ____/_____ ___   ____ _ / /_   / ___/ __  __ _____ _____ ___   _____ _____
 / / __ / ___// _ \ / __ `// __/   \__ \ / / / // ___// ___// _ \ / ___// ___/
/ /_/ // /   /  __// /_/ // /_    ___/ // /_/ // /__ / /__ /  __/(__  )(__  ) 
\____//_/    \___/ \__,_/ \__/   /____/ \__,_/ \___/ \___/ \___//____//____/  
                                                                              

""")